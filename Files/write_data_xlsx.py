from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# ws['A1'] = "Software tester"
# ws['B1'] = "Sharmila"
# testdata = [["Name", "City"], ["Sharmila", "Madurai"]]
# for data in testdata:
#     ws.append(data)
for i in range(1, 6):
    for j in range(1, 6):
        ws.cell(row=i, column=j).value = i + j


wb.save("createexcel.xlsx")
