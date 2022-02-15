from openpyxl import workbook, load_workbook
wb = load_workbook(filename="createexcel.xlsx")
sh = wb.active
print(sh['A1'].value)

row_ct = sh.max_row
col_ct = sh.max_column

print(row_ct)
print(col_ct)

for i in range(1, row_ct+1):
    for j in range(1, col_ct+1):
       print(sh.cell(row=i, column=j).value, end='')
       print('\n')
